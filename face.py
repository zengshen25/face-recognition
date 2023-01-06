import cv2
import numpy as np
import os
import shutil
import threading
import tkinter as tk
from PIL import Image, ImageTk
import openpyxl

# 首先读取config文件，第一行代表当前已经储存的人名个数，接下来每一行是（id，name）标签和对应的人名
# 字典里存的是id——name键值对
id_dict = {}
# 已经被识别有用户名的人脸个数
Total_face_num = 888


# 将config文件内的信息读入到字典中
def init():
    f = open('./config.txt')
    global Total_face_num

    Total_face_num = len(f.readlines())
    f = open('./config.txt')
    for i in range(int(Total_face_num)):
        line = f.readline()
        id_name = line.split(' ')

        id_dict[int(id_name[0])] = id_name[1]
    f.close()


init()

# 加载OpenCV人脸检测分类器Haar
face_cascade = cv2.CascadeClassifier(r"./resources/haarcascade_frontalface_default.xml")

# 准备好识别方法LBPH方法
recognizer = cv2.face.LBPHFaceRecognizer_create()

# 打开标号为0的摄像头
camera = cv2.VideoCapture(0)

# 从摄像头读取照片
success, img = camera.read()
W_size = 0.1 * camera.get(3)
H_size = 0.1 * camera.get(4)

# 标志系统状态的量 0表示无子线程在运行 1表示正在刷脸 2表示正在录入新面孔。
# 相当于mutex锁，用于线程同步
system_state_lock = 0


def Train_new_face():
    print("\n正在训练中")
    # cv2.destroyAllWindows()
    path = 'data'

    # 初始化识别的方法
    recognizer = cv2.face.LBPHFaceRecognizer_create()

    # 调用函数并将数据喂给识别器训练
    faces, ids = get_images_and_labels(path)
    print('本次用于训练的识别码为:')  # 调试信息
    print(ids)  # 输出识别码

    # 训练模型  #将输入的所有图片转成四维数组
    recognizer.train(faces, np.array(ids))
    # 保存模型
    print(Total_face_num)

    yml = 'face-data/' + str(Total_face_num) + ".yml"
    rec_f = open(yml, "w+")
    rec_f.close()
    recognizer.save(yml)


def get_new_face():
    print("正在从摄像头录入您的人脸信息 \n")

    # 存在目录data就清空，不存在就创建，确保最后存在空的data目录
    filepath = "data"
    if not os.path.exists(filepath):
        os.mkdir(filepath)
    else:
        shutil.rmtree(filepath)
        os.mkdir(filepath)

    # 已经获得的样本数
    sample_num = 0
    # 从摄像头读取图片
    while True:

        global success
        # 因为要显示在可视化的控件内，所以要用全局的
        global img
        success, img = camera.read()

        # 转为灰度图片
        if success is True:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        else:
            break

        # 检测人脸，将每一帧摄像头记录的数据带入OpenCv中，让Classifier判断人脸
        # 其中gray为要检测的灰度图像，1.3为每次图像尺寸减小的比例，5为minNeighbors
        face_detector = face_cascade
        faces = face_detector.detectMultiScale(gray, 1.3, 5)

        # 框选人脸，for循环保证一个能检测的实时动态视频流
        for (x, y, w, h) in faces:
            # xy为左上角的坐标,w为宽，h为高，用rectangle为人脸标记画框
            cv2.rectangle(img, (x, y), (x + w, y + w), (255, 0, 0))
            # 样本数加1
            sample_num += 1
            # 保存图像，把灰度图片看成二维数组来检测人脸区域，这里是保存在data缓冲文件夹内
            T = Total_face_num
            cv2.imwrite("./data/User." + str(T) + '.' + str(sample_num) + '.jpg', gray[y:y + h, x:x + w])

        # 表示摄像头拍摄取样的数量,越多效果越好，但获取以及训练的越慢
        pictur_num = 30

        cv2.waitKey(1)
        if sample_num > pictur_num:
            break
        # 控制台内输出进度条
        else:
            l = int(sample_num / pictur_num * 50)
            r = int((pictur_num - sample_num) / pictur_num * 50)
            print("\r" + "%{:.1f}".format(sample_num / pictur_num * 100) + "=" * l + "->" + "_" * r, end="")
            # 控件可视化进度信息
            var.set("%{:.1f}".format(sample_num / pictur_num * 100))
            # 刷新控件以实时显示进度
            window.update()


# 创建一个函数，用于从数据集文件夹中获取训练图片,并获取id
# 注意图片的命名格式为User.id.sampleNum
def get_images_and_labels(path):
    image_paths = [os.path.join(path, f) for f in os.listdir(path)]
    # 新建连个list用于存放
    face_samples = []
    ids = []

    # 遍历图片路径，导入图片和id添加到list中
    for image_path in image_paths:

        # 通过图片路径将其转换为灰度图片
        img = Image.open(image_path).convert('L')

        # 将图片转化为数组
        img_np = np.array(img, 'uint8')

        if os.path.split(image_path)[-1].split(".")[-1] != 'jpg':
            continue

        # 为了获取id，将图片和路径分裂并获取
        id = int(os.path.split(image_path)[-1].split(".")[1])

        # 调用熟悉的人脸分类器
        detector = cv2.CascadeClassifier(r"./resources/haarcascade_frontalface_default.xml")

        faces = detector.detectMultiScale(img_np)

        # 将获取的图片和id添加到list中
        for (x, y, w, h) in faces:
            face_samples.append(img_np[y:y + h, x:x + w])
            ids.append(id)
    return face_samples, ids


def write_config():
    print("新的人脸训练结束")
    f = open('config.txt', "a")
    T = Total_face_num
    data = openpyxl.load_workbook('user.xlsx')
    sheet_names = data.sheetnames
    table = data[sheet_names[0]]
    table = data.active
    # 获得行数
    nrows = table.max_row
    # 获得列数
    ncolumns = table.max_column
    table.cell(nrows + 1, 1).value = T
    name = ""
    for i in range(2, ncolumns + 1):
        table.cell(nrows + 1, i).value = input("输入" + str(table.cell(1, i).value) + ": ")
        if i == 2:
            name = str(table.cell(nrows + 1, i).value)
    f.write(str(T) + " " + name + " \n")
    f.close()
    id_dict[T] = name
    data.save('user.xlsx')
    # 这里修改文件的方式是先读入内存，然后修改内存中的数据，最后写回文件
    f = open('config.txt', 'r+')
    flist = f.readlines()
    f.close()
    f = open('config.txt', 'w+')
    f.writelines(flist)
    f.close()

def face_swiping():
    # 使用之前训练好的模型
    for i in range(Total_face_num):  # 每个识别器都要用
        i += 1
        yml = './face-data/' + str(i) + ".yml"
        print(yml)
        recognizer.read(yml)

        ave_poss = 0
        # 每个识别器扫描十遍
        for times in range(10):
            times += 1
            cur_poss = 0
            global success
            global img

            global system_state_lock
            while system_state_lock == 2:  # 如果正在录入新面孔就阻塞
                print("\r刷脸被录入面容阻塞", end="")
                pass

            success, img = camera.read()
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            # 识别人脸
            faces = face_cascade.detectMultiScale(
                gray,
                scaleFactor=1.2,
                minNeighbors=5,
                minSize=(int(W_size), int(H_size))
            )
            # 进行校验
            for (x, y, w, h) in faces:

                while system_state_lock == 2:  # 如果正在录入新面孔就阻塞
                    print("\r刷脸被录入面容阻塞", end="")
                    pass
                # 这里c
                cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
                # 调用分类器的预测函数，接收返回值标签和置信度
                idnum, confidence = recognizer.predict(gray[y:y + h, x:x + w])
                conf = confidence
                # 计算出一个检验结果
                if confidence < 100:  # 可以识别出已经训练的对象——直接输出姓名在屏幕上
                    if idnum in id_dict:
                        user_name = id_dict[idnum]
                    else:
                        user_name = "Untagged user:" + str(idnum)
                    confidence = "{0}%", format(round(100 - confidence))
                else:  # 无法识别此对象，那么就开始训练
                    user_name = "unknown"
                # 加载一个字体用于输出识别对象的信息
                font = cv2.FONT_HERSHEY_SIMPLEX
                # 输出检验结果以及用户名
                cv2.putText(img, str(user_name), (x + 5, y - 5), font, 1, (0, 0, 255), 1)
                cv2.putText(img, str(confidence), (x + 5, y + h - 5), font, 1, (0, 0, 0), 1)
                if 15 > conf > 0:
                    # 表示可以识别
                    cur_poss = 1
                elif 60 > conf > 35:
                    # 表示可以识别
                    cur_poss = 1
                else:
                    # 表示不可以识别
                    cur_poss = 0
            k = cv2.waitKey(1)
            if k == 27:
                cv2.destroyAllWindows()
                break
            ave_poss += cur_poss
        # 有一半以上识别说明可行则返回
        if ave_poss >= 5:
            return i
    # 全部过一遍还没识别出说明无法识别
    return 0



def scan_face_recognition_thread():
    var.set('刷脸')
    ans = face_swiping()
    if ans == 0:
        print("最终结果：无法识别")
        var.set("最终结果：无法识别")
    else:
        ans_name = "最终结果：" + id_dict[ans]
        print(ans)
        wb = openpyxl.load_workbook('user.xlsx')
        sheet = wb['Sheet1']
        ans_name.strip()
        print("用户信息为：")
        content1 = ["编号: ", "姓名: ", "部门: ", "职位: ", "年龄: "]
        i = 0
        content = ""
        for k in range(1, sheet.max_row + 1):
            if sheet.cell(row=k, column=1).value == ans:
                for c in range(1, sheet.max_column + 1):
                    content = content + content1[i] + '  ' + str(sheet.cell(row=k, column=c).value) + "\n"
                    i = i + 1
        var.set(ans_name)
        var_content.set(content)
        global photo2
        photo2 = tk.PhotoImage(file="user-img/" + str(ans) + ".png")
        # 把图片整合到标签类中
        img_label = tk.Label(window, image=photo2, width=180, height=180)
        # 自动对齐
        img_label.place(x=568, y=140)
    global system_state_lock
    # 修改system_state_lock,释放资源
    system_state_lock = 0


def scan_face_recognition():
    # 重置上次留下的信息
    var_content.set('')
    photo2 = tk.PhotoImage(file="user-img/" + "测试.png")
    img_label = tk.Label(window, image=photo2, width=180, height=180)
    img_label.place(x=568, y=140)
    global system_state_lock
    if system_state_lock == 1:
        print("阻塞，因为正在刷脸")
        return 0
    elif system_state_lock == 2:  # 如果正在录入新面孔就阻塞
        print("\n刷脸被录入面容阻塞\n" "")
        return 0
    system_state_lock = 1
    p = threading.Thread(target=scan_face_recognition_thread)
    # 把线程P设置为守护线程 若主线程退出 P也跟着退出
    p.setDaemon(True)
    p.start()


def enter_face_recognition_thread():
    var.set('录入')
    cv2.destroyAllWindows()
    global Total_face_num
    Total_face_num += 1
    # 采集新人脸
    get_new_face()
    print("采集完毕，开始训练")
    # 采集完就可以解开锁
    global system_state_lock
    # print("锁被释放0")
    system_state_lock = 0
    # 训练采集到的新人脸
    Train_new_face()
    # 修改配置文件
    write_config()


# 点击录入人脸的按钮触发的函数
def enter_face_recognition():
    global system_state_lock
    if system_state_lock == 2:
        print("阻塞，因为正在录入面容")
        return 0
    else:
        system_state_lock = 2  # 修改system_state_lock
        print("改为2", end="")
        # print("当前锁的值为：" + str(system_state_lock))

    p = threading.Thread(target=enter_face_recognition_thread)
    # 把线程P设置为守护线程 若主线程退出 P也跟着退出
    p.setDaemon(True)
    p.start()


# 退出按钮
def face_button_exit():
    exit()



# 用于在label内动态展示摄像头内容（摄像头嵌入控件）
def video_loop():
    # success, img = camera.read()  # 从摄像头读取照片
    global success
    global img
    if success:
        cv2.waitKey(1)
        # 转换颜色从BGR到RGBA
        cv2image = cv2.cvtColor(img, cv2.COLOR_BGR2RGBA)
        # 将图像转换成Image对象
        current_image = Image.fromarray(cv2image)
        img_tk = ImageTk.PhotoImage(image=current_image)
        panel.img_tk = img_tk
        panel.config(image=img_tk)
        window.after(1, video_loop)


window = tk.Tk()
# 窗口标题
window.title('人脸识别考勤算法设计与实现')
window.geometry('1000x610')

# 在图形界面上设定标签，类似于一个提示窗口的作用
var = tk.StringVar()
label_control = tk.Label(window, textvariable=var, bg='#dee3e9', fg='black', font=('Arial', 12), width=50, height=4)
label_control.pack()

# 这是我自己写的列表控件样式
var_content = tk.StringVar()
button_a1 = tk.Label(window, justify='left', textvariable=var_content, font=('Arial', 16), width=70, height=20)
button_a1.place(x=220, y=180)

# 在窗口界面设置放置Button按键并绑定处理函数
button_a1 = tk.Button(window, text='开始刷脸', font=('Arial', 12), width=10, height=2, command=scan_face_recognition)
button_a1.place(x=800, y=200)

button_b1 = tk.Button(window, text='录入人脸', font=('Arial', 12), width=10, height=2, command=enter_face_recognition)
button_b1.place(x=800, y=300)

button_b1 = tk.Button(window, text='退出', font=('Arial', 12), width=10, height=2, command=face_button_exit)
button_b1.place(x=800, y=400)

# 摄像头模块大小
panel = tk.Label(window, width=500, height=350)
# 摄像头模块的位置
panel.place(x=10, y=140)
window.config(cursor="arrow")

video_loop()
#  窗口循环，用于显示
window.mainloop()
